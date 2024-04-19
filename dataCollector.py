from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import sqlite3
import snap7
from snap7 import util

def snapRun(device_letter, ip_address):
    counterDword = [0]
    counterReal = [16, 52, 88, 124, 196, 232, 268, 304, 376, 340, 160]
    client = snap7.client.Client()
    client.connect(ip_address, 0, 1)
    client.get_connected()

    values = [device_letter, datetime.now().strftime("%Y-%m-%d %H:%M:%S")]

    for countDword in counterDword:
        db_Dword = client.db_read(60, countDword, 4)
        d = util.get_dint(db_Dword, 0)
        d_formatted = "{:.2f}".format(d)
        values.append(d_formatted)

    for countReal in counterReal:
        db_real = client.db_read(88, countReal, 4)
        t = util.get_real(db_real, 0)
        t_formatted = "{:.2f}".format(t)
        values.append(t_formatted)

    client.disconnect()

    return values

def find_empty_column(ws, row):
    col = 3
    while ws.cell(row=row, column=col).value is not None:
        col += 1
    return get_column_letter(col)

def create_or_load_workbook():
    current_date = datetime.now().strftime("%Y-%m-%d")
    target_filename = f"{current_date}_Daily_Check_list.xlsx"
    if os.path.exists(target_filename):
        return load_workbook(target_filename)
    else:
        result_wb = load_workbook('result.xlsx')
        result_ws = result_wb.active
        result_wb.save(target_filename)
        return load_workbook(target_filename)

try:
    conn = sqlite3.connect('mysite/db.sqlite3')
    cur = conn.cursor()

    ip_addresses = {'A': '192.168.1.15', 'B': '192.168.1.12'}

    for device_letter, ip_address in ip_addresses.items():
        
        db_real = client.db_read(88, 124, 4)
        if db_real < 400:
            wb = create_or_load_workbook()
            snap_data = snapRun(device_letter, ip_address)
            new_sheet_name = snap_data[0]

            if new_sheet_name not in wb.sheetnames:
                print(f"Sheet '{new_sheet_name}' not found, creating a new sheet...")
                new_sheet = wb.create_sheet(new_sheet_name, index=0)
            else:
                print(f"Sheet '{new_sheet_name}' found, using existing sheet...")
                new_sheet = wb[new_sheet_name]

            empty_column = find_empty_column(new_sheet, row=6)
            last_column_index = ord(empty_column) - 64

            for idx, param_value in enumerate(snap_data[1:], start=6):
                new_sheet.cell(row=idx, column=last_column_index, value=param_value)

            wb.save(f"{datetime.now().strftime('%Y-%m-%d')}_Daily_Check_list.xlsx")
            print("Saved")

        data = snapRun(device_letter, ip_address)
        cur.execute('''INSERT INTO app_parameter(
            comp_number, time, hour_meter, inlet_pressure, stage_1_pressure, discharge_pressure, 
            comp_oil_temp, compressor_oil_presure, gas_detector, motor_current, st_1_1_temp, 
            st_1_2_temp, st_2_1_temp, st_2_2_temp) VALUES
            (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);''', tuple(data))

    conn.commit()
    conn.close()
    print("Data inserted and Excel updated successfully")
except Exception as e:
    print("Error:", e)
