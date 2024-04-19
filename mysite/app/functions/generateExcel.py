from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
from .snapData import snapRun  

def find_empty_column(ws, row):
    col = 3
    while ws.cell(row=row, column=col).value is not None:
        col += 1
    return get_column_letter(col)

def create_or_load_workbook():
    current_date = datetime.now().strftime("%Y-%m-%d")
    target_filename = f"{current_date}_Daily_Check_list.xlsx"

    if os.path.exists(target_filename):
        return load_workbook(target_filename)
    else:
        result_wb = load_workbook('result.xlsx')
        result_ws = result_wb.active
        result_wb.save(target_filename) 
        return load_workbook(target_filename) 

def FillWorkBook():
    wb = create_or_load_workbook()
    ws = wb.active 
    snap_data = snapRun()


    col_to_insert = find_empty_column(ws, row=6)
    for idx, param_value in enumerate(snap_data.values(), start=6):
        ws[f"{col_to_insert}{idx}"] = param_value

    wb.save(f"{datetime.now().strftime('%Y-%m-%d')}_Daily_Check_list.xlsx")
    print("Saved")

FillWorkBook()